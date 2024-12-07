from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# Criação da planilha
wb = Workbook()
ws = wb.active
ws.title = "Planilha Inteligente"

# Títulos das colunas
headers = ["Produto", "Quantidade", "Preço Unitário", "Total"]
ws.append(headers)

# Estilizando os títulos
for col in ws.iter_cols(min_row=1, max_row=1, max_col=len(headers)):
    for cell in col:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

# Função para entrada de dados
def adicionar_produto():
    while True:
        produto = input("Digite o nome do produto (ou 'sair' para finalizar): ")
        if produto.lower() == "sair":
            break
        
        quantidade = int(input("Digite a quantidade: "))
        preco_unitario = float(input("Digite o preço unitário: R$ "))
        
        # Adicionando dados à planilha
        total = quantidade * preco_unitario
        ws.append([produto, quantidade, preco_unitario, total])
        print(f"Produto '{produto}' adicionado com sucesso!\n")

# Função para calcular o total geral
def calcular_total_geral():
    total_geral = "=SUM(D2:D{})".format(ws.max_row)
    ws[f"C{ws.max_row + 1}"] = "Total Geral:"
    ws[f"C{ws.max_row}"].font = Font(bold=True)
    ws[f"C{ws.max_row}"].alignment = Alignment(horizontal="right")
    ws[f"D{ws.max_row}"] = total_geral
    ws[f"D{ws.max_row}"].font = Font(bold=True)

# Adicionando produtos
adicionar_produto()

# Calculando o total geral
calcular_total_geral()

# Salvando o arquivo
nome_arquivo = "planilha_inteligente.xlsx"
wb.save(nome_arquivo)
print(f"\nPlanilha '{nome_arquivo}' criada com sucesso!")
